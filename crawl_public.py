"""
豆瓣数据备份工具 - 针对特定用户
无需登录，爬取公开数据
"""
import json
import os
import re
import time
import sys
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
from backup_metadata import build_metadata, merge_metadata, metadata_rows
from diagnostics import classify_response
from excel_safety import sanitize_excel_value

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

BASE_URL = 'https://www.douban.com'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
}

USER_ID = None  # 由命令行参数或交互输入设置
SESSION = requests.Session()
SESSION.headers.update(HEADERS)

DEFAULT_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'backup')
OUTPUT_DIR = DEFAULT_OUTPUT_DIR
os.makedirs(OUTPUT_DIR, exist_ok=True)
DEFAULT_CATEGORIES = ['movies', 'books', 'music', 'games']


def get_comment(item):
    """Return the user comment from a collection item, when present."""
    comment_tag = item.select_one('.comment')
    return comment_tag.get_text(' ', strip=True) if comment_tag else ''


def crawl_movies():
    """爬取电影数据"""
    print("\n[电影] 爬取电影数据...")
    all_movies = {'wish': [], 'collect': [], 'do': []}

    collections = [
        ('collect', '已看'),
        ('do', '在看'),
        ('wish', '想看')
    ]

    for coll_type, coll_name in collections:
        print(f"\n  爬取 {coll_name} 的电影...")
        url = f"https://movie.douban.com/people/{USER_ID}/{coll_type}"
        page = 0
        total = 0

        while True:
            page_url = f"{url}?start={page * 15}&sort=time"
            try:
                response = SESSION.get(page_url, timeout=30)
                status, _, message = classify_response(response)
                if status != 'ok':
                    print(f"    [WARN] {message}")
                    break

                soup = BeautifulSoup(response.text, 'html.parser')
                items = soup.find_all('div', class_='item')

                if not items:
                    break

                for item in items:
                    try:
                        movie = parse_movie_item(item)
                        if movie:
                            all_movies[coll_type].append(movie)
                            total += 1
                    except Exception as e:
                        continue

                print(f"    第{page + 1}页: 获取 {len(items)} 条")

                if len(items) < 15:
                    break

                page += 1
                time.sleep(1)

            except Exception as e:
                print(f"    错误: {e}")
                break

        print(f"  {coll_name}: {total} 部")

    return all_movies


def parse_movie_item(item):
    """解析电影条目"""
    try:
        info = item.find('div', class_='info')
        if not info:
            return None

        title_tag = info.find('a', title=True)
        title = title_tag.get('title', '') if title_tag else ''

        if not title:
            ul = info.find('ul')
            if ul:
                title_tag = ul.find('span', class_='title')
                title = title_tag.get_text(strip=True) if title_tag else ''

        url = title_tag.get('href', '') if title_tag else ''

        subject_id = ''
        if url:
            match = re.search(r'subject/(\d+)', url)
            if match:
                subject_id = match.group(1)

        cover_img = item.find('img')
        cover = cover_img.get('src', '') if cover_img else ''

        rating_tag = item.find('span', class_=lambda x: x and 'rating' in x)
        rating = ''
        if rating_tag:
            rating_class = rating_tag.get('class', [])
            for cls in rating_class:
                if 'rating' in cls and 't' in cls:
                    match = re.search(r'rating(\d+)', cls)
                    if match:
                        rating = match.group(1)
                    break

        date_tag = item.find('span', class_='date')
        date = date_tag.get_text(strip=True) if date_tag else ''

        comment = get_comment(item)

        return {
            'douban_id': subject_id,
            'title': title,
            'cover': cover,
            'rating': rating,
            'date': date,
            'comment': comment
        }
    except Exception:
        return None


def crawl_books():
    """爬取书籍数据"""
    print("\n[书籍] 爬取书籍数据...")
    all_books = {'wish': [], 'collect': [], 'reading': []}

    collections = [
        ('collect', '已读'),
        ('reading', '在读'),
        ('wish', '想读')
    ]

    for coll_type, coll_name in collections:
        print(f"\n  爬取 {coll_name} 的书籍...")
        url = f"https://book.douban.com/people/{USER_ID}/{coll_type}"
        page = 0
        total = 0

        while True:
            page_url = f"{url}?start={page * 15}"
            try:
                response = SESSION.get(page_url, timeout=30)
                status, _, message = classify_response(response)
                if status != 'ok':
                    print(f"    [WARN] {message}")
                    break

                soup = BeautifulSoup(response.text, 'html.parser')
                items = soup.find_all('li', class_='subject-item')

                if not items:
                    items = soup.find_all('div', class_='item') # old style/fallback
                
                if not items:
                    items = soup.find_all('li', class_='item') # another variant

                if not items:
                    break

                for item in items:
                    try:
                        book = parse_book_item(item)
                        if book:
                            all_books[coll_type].append(book)
                            total += 1
                    except Exception:
                        continue

                print(f"    第{page + 1}页: 获取 {len(items)} 条")

                if len(items) < 15:
                    break

                page += 1
                time.sleep(1)

            except Exception as e:
                print(f"    错误: {e}")
                break

        print(f"  {coll_name}: {total} 本")

    return all_books


def parse_book_item(item):
    """解析书籍条目"""
    try:
        # Title
        # Try new style: div.info h2 a
        title = ''
        url = ''
        
        info = item.find('div', class_='info')
        if info:
            h2 = info.find('h2')
            if h2:
                a = h2.find('a')
                if a:
                    title = a.get_text(strip=True)
                    url = a.get('href', '')
        
        # Fallback to old style
        if not title:
            title_tag = item.find('span', class_='title')
            if title_tag:
                 title = title_tag.get_text(strip=True)
                 a = item.find('a', href=True)
                 if a: url = a.get('href', '')

        subject_id = ''
        if 'subject' in url:
            match = re.search(r'subject/(\d+)', url)
            if match:
                subject_id = match.group(1)

        img_tag = item.find('img')
        cover = img_tag.get('src', '') if img_tag else ''

        # Author/Pub
        author = ''
        pub = item.find('div', class_='pub')
        if pub:
            author = pub.get_text(strip=True)
        else:
            author_tag = item.find('span', class_='author')
            if author_tag: author = author_tag.get_text(strip=True)

        # Rating
        rating = ''
        rating_tag = item.find('span', class_=lambda x: x and 'rating' in x)
        if rating_tag:
            rating_class = rating_tag.get('class', [])
            for cls in rating_class:
                if 'rating' in cls and 't' in cls:
                    match = re.search(r'rating(\d+)', cls)
                    if match:
                        rating = match.group(1)
                    break
        
        comment = get_comment(item)

        return {
            'douban_id': subject_id,
            'title': title,
            'author': author,
            'cover': cover,
            'rating': rating,
            'comment': comment
        }
    except Exception:
        return None


def crawl_music():
    """爬取音乐数据"""
    print("\n[音乐] 爬取音乐数据...")
    all_music = {'wish': [], 'collect': [], 'do': []}

    collections = [
        ('collect', '听过'),
        ('do', '在听'),
        ('wish', '想听')
    ]

    for coll_type, coll_name in collections:
        print(f"\n  爬取 {coll_name} 的音乐...")
        url = f"https://music.douban.com/people/{USER_ID}/{coll_type}"
        page = 0
        total = 0

        while True:
            page_url = f"{url}?start={page * 15}&sort=time"
            try:
                response = SESSION.get(page_url, timeout=30)
                status, _, message = classify_response(response)
                if status != 'ok':
                    print(f"    [WARN] {message}")
                    break

                soup = BeautifulSoup(response.text, 'html.parser')
                items = soup.select('div.item')
                
                if not items:
                    break

                for item in items:
                    try:
                        music = parse_music_item(item)
                        if music:
                            all_music[coll_type].append(music)
                            total += 1
                    except Exception:
                        continue

                print(f"    第{page + 1}页: 获取 {len(items)} 条")

                if len(items) < 15:
                    break

                page += 1
                time.sleep(1)

            except Exception as e:
                print(f"    错误: {e}")
                break

        print(f"  {coll_name}: {total} 张")

    return all_music


def parse_music_item(item):
    """解析音乐条目"""
    try:
        info = item.select_one('.info')
        if not info: 
            return None
            
        title_tag = info.select_one('a em') or info.select_one('a')
        title = title_tag.get_text(strip=True) if title_tag else ''
        
        a_tag = info.select_one('a')
        url = a_tag.get('href', '') if a_tag else ''
        subject_id = ''
        if 'subject' in url:
            match = re.search(r'subject/(\d+)', url)
            if match:
                subject_id = match.group(1)

        img_tag = item.select_one('img')
        cover = img_tag.get('src', '') if img_tag else ''

        rating_tag = item.select_one('span[class^="rating"]')
        rating = ''
        if rating_tag:
            rating_class = rating_tag.get('class', [])
            for cls in rating_class:
                if 'rating' in cls and 't' in cls:
                    match = re.search(r'rating(\d+)', cls)
                    if match:
                        rating = match.group(1)
                    break
        
        intro_tag = info.select_one('li.intro')
        intro = intro_tag.get_text(strip=True) if intro_tag else ''
        artist = intro.split('/')[0].strip() if intro else ''

        comment = get_comment(item)

        return {
            'douban_id': subject_id,
            'title': title,
            'artist': artist,
            'cover': cover,
            'rating': rating,
            'info': intro,
            'comment': comment
        }
    except Exception:
        return None


def crawl_games():
    """爬取游戏数据"""
    print("\n[游戏] 爬取游戏数据...")
    all_games = {'wish': [], 'collect': [], 'do': []}

    collections = [
        ('collect', '玩过'),
        ('do', '在玩'),
        ('wish', '想玩')
    ]

    for coll_type, coll_name in collections:
        print(f"\n  爬取 {coll_name} 的游戏...")
        url = f"https://www.douban.com/people/{USER_ID}/games" 
        page = 0
        total = 0

        while True:
            page_url = f"{url}?action={coll_type}&start={page * 15}"
            try:
                response = SESSION.get(page_url, timeout=30)
                status, _, message = classify_response(response)
                if status != 'ok':
                    print(f"    [WARN] {message}")
                    break

                soup = BeautifulSoup(response.text, 'html.parser')
                items = soup.select('div.common-item')
                
                if not items:
                    break

                for item in items:
                    try:
                        game = parse_game_item(item)
                        if game:
                            all_games[coll_type].append(game)
                            total += 1
                    except Exception:
                        continue

                print(f"    第{page + 1}页: 获取 {len(items)} 条")

                if len(items) < 15:
                    break

                page += 1
                time.sleep(1)

            except Exception as e:
                print(f"    错误: {e}")
                break

        print(f"  {coll_name}: {total} 个")

    return all_games


def parse_game_item(item):
    """解析游戏条目"""
    try:
        title_tag = item.select_one('.title a')
        title = title_tag.get_text(strip=True) if title_tag else ''
        
        url = title_tag.get('href', '') if title_tag else ''
        subject_id = ''
        match = re.search(r'(?:subject|game)/(\d+)', url)
        if match:
            subject_id = match.group(1)

        img_tag = item.select_one('img')
        cover = img_tag.get('src', '') if img_tag else ''

        rating = ''
        rating_tag = item.select_one('span[class^="allstar"]') or item.select_one('span.rating-star')
        if rating_tag:
            classes = rating_tag.get('class', [])
            for cls in classes:
                if cls.startswith('allstar'):
                    rating = cls.replace('allstar', '')
                    if len(rating) == 2:
                        rating = str(int(rating) // 10)
                    break
            
            if not rating and rating_tag.get('title'):
                rating = rating_tag.get('title')

        desc_tag = item.select_one('.desc')
        desc = desc_tag.get_text(strip=True) if desc_tag else ''
        
        date = ''
        if desc:
             parts = desc.split('/')
             if parts:
                 date = parts[0].strip()

        comment = get_comment(item)

        return {
            'douban_id': subject_id,
            'title': title,
            'cover': cover,
            'rating': rating,
            'date': date,
            'info': desc,
            'comment': comment
        }
    except Exception:
        return None


def save_json(data, filename, metadata=None):
    """保存JSON文件"""
    filepath = os.path.join(OUTPUT_DIR, f"{filename}.json")
    payload = {
        'metadata': merge_metadata(metadata),
        'data': data,
    }
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"\n[OK] 已保存: {filepath}")
    return filepath


def save_excel(data, filename, metadata=None):
    """保存Excel文件"""
    filepath = os.path.join(OUTPUT_DIR, f"{filename}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = '数据'

    metadata_sheet = wb.create_sheet('元数据', 0)
    metadata_sheet.column_dimensions['A'].width = 24
    metadata_sheet.column_dimensions['B'].width = 48
    for row_index, (key, value) in enumerate(metadata_rows(metadata), 1):
        metadata_sheet.cell(row=row_index, column=1, value=key)
        metadata_sheet.cell(
            row=row_index,
            column=2,
            value=sanitize_excel_value(value),
        )
    wb.active = 1

    ws.append(['类型', '收藏状态', '标题', '豆瓣ID', '评分', '评语', '作者/音乐人/日期', '日期/封面/详情'])

    for category, collections in data.items():
        for coll, items in collections.items():
            for item in items:
                # Normalize fields for different types
                field1 = item.get('author') or item.get('artist') or item.get('year') or item.get('date') or ''
                field2 = item.get('date') or item.get('cover') or item.get('info') or ''
                
                row = [
                    sanitize_excel_value(category),
                    sanitize_excel_value(coll),
                    sanitize_excel_value(item.get('title', '')),
                    sanitize_excel_value(item.get('douban_id', '')),
                    sanitize_excel_value(item.get('rating', '')),
                    sanitize_excel_value(item.get('comment', '')),
                    sanitize_excel_value(field1),
                    sanitize_excel_value(field2)
                ]
                ws.append(row)

    wb.save(filepath)
    print(f"[OK] 已保存: {filepath}")
    return filepath


def run_public_backup(user_id, categories=None, output_dir=None):
    global USER_ID, OUTPUT_DIR

    USER_ID = user_id
    OUTPUT_DIR = output_dir or DEFAULT_OUTPUT_DIR
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    categories = list(categories or DEFAULT_CATEGORIES)
    metadata = build_metadata(
        backup_mode='public',
        selected_categories=categories,
        user_id=user_id,
        output_dir=OUTPUT_DIR,
    )

    print("=" * 50)
    print("[豆瓣数据备份工具 - 公开数据]")
    print("=" * 50)
    print(f"\n用户: {USER_ID}")
    print(f"时间: {metadata['generated_at']}")

    timestamp = datetime.now().astimezone().strftime('%Y%m%d_%H%M%S')
    all_data = {}

    crawlers = {
        'movies': crawl_movies,
        'books': crawl_books,
        'music': crawl_music,
        'games': crawl_games,
    }

    try:
        for category in categories:
            category_data = crawlers[category]()
            all_data[category] = category_data
            save_json(
                category_data,
                f"{category}_{timestamp}",
                metadata=build_metadata(
                    backup_mode='public',
                    selected_categories=[category],
                    user_id=user_id,
                    output_dir=OUTPUT_DIR,
                ),
            )

        save_json(all_data, f"douban_backup_{timestamp}", metadata=metadata)
        save_excel(all_data, f"douban_backup_{timestamp}", metadata=metadata)

        print("\n" + "=" * 50)
        print("[备份统计]")
        print("=" * 50)
        for category in categories:
            total = sum(len(v) for v in all_data.get(category, {}).values())
            label = {
                'movies': '电影',
                'books': '书籍',
                'music': '音乐',
                'games': '游戏',
            }[category]
            unit = {
                'movies': '部',
                'books': '本',
                'music': '张',
                'games': '个',
            }[category]
            print(f"  {label}: {total} {unit}")
        print(f"\n文件保存在: {OUTPUT_DIR}")
        return all_data
    except KeyboardInterrupt:
        print("\n\n用户中断，保存已获取的数据...")
        if all_data:
            save_json(all_data, f"douban_backup_interrupted_{timestamp}", metadata=metadata)
            save_excel(all_data, f"douban_backup_interrupted_{timestamp}", metadata=metadata)
        return all_data


def main():
    global USER_ID

    if len(sys.argv) > 1:
        USER_ID = sys.argv[1]
    else:
        USER_ID = input("请输入豆瓣用户ID: ").strip()
        if not USER_ID:
            print("用户ID不能为空!")
            print("用法: python crawl_public.py <用户ID>")
            return

    print("=" * 50)
    print("[豆瓣数据备份工具]")
    print("=" * 50)
    run_public_backup(USER_ID)


if __name__ == '__main__':
    main()
