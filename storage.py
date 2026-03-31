"""
数据存储模块
支持保存为JSON和Excel格式（美化版）
"""
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import DATA_DIR


# 星级显示
def _rating_to_stars(rating):
    """将评分数字转为星级符号"""
    try:
        n = int(rating)
        if n < 1 or n > 5:
            return str(rating)
        return '\u2605' * n + '\u2606' * (5 - n)
    except (ValueError, TypeError):
        return ''


# 豆瓣链接
_LINK_PREFIX = {
    'movie': 'https://movie.douban.com/subject/',
    'book': 'https://book.douban.com/subject/',
    'music': 'https://music.douban.com/subject/',
    'game': 'https://www.douban.com/game/',
}

# 各类别的列定义: (字段key, 中文表头, 列宽)
_COLUMNS = {
    'movies': [
        ('_index', '序号', 6),
        ('title', '标题', 30),
        ('rating', '我的评分', 12),
        ('comment', '我的评语', 40),
        ('tags', '标签', 20),
        ('date', '标记日期', 14),
        ('_link', '豆瓣链接', 40),
    ],
    'books': [
        ('_index', '序号', 6),
        ('title', '标题', 30),
        ('rating', '我的评分', 12),
        ('comment', '我的评语', 40),
        ('author', '作者/出版', 30),
        ('date', '标记日期', 14),
        ('_link', '豆瓣链接', 40),
    ],
    'music': [
        ('_index', '序号', 6),
        ('title', '标题', 30),
        ('rating', '我的评分', 12),
        ('comment', '我的评语', 40),
        ('artist', '艺术家', 20),
        ('info', '简介', 30),
        ('_link', '豆瓣链接', 40),
    ],
    'games': [
        ('_index', '序号', 6),
        ('title', '标题', 30),
        ('rating', '我的评分', 12),
        ('comment', '我的评语', 40),
        ('info', '简介', 30),
        ('date', '标记日期', 14),
        ('_link', '豆瓣链接', 40),
    ],
}

# 类别中文名 & 量词
_CATEGORY_CN = {
    'movies': ('电影', '部'),
    'books': ('书籍', '本'),
    'music': ('音乐', '张'),
    'games': ('游戏', '个'),
}

# 状态排列顺序（已完成优先）& 颜色
_STATUS_ORDER = {
    'movies': [
        ('collect', '看过', '4CAF50'),
        ('do', '在看', '2196F3'),
        ('wish', '想看', 'FF9800'),
    ],
    'books': [
        ('collect', '已读', '4CAF50'),
        ('reading', '在读', '2196F3'),
        ('wish', '想读', 'FF9800'),
    ],
    'music': [
        ('collect', '听过', '4CAF50'),
        ('do', '在听', '2196F3'),
        ('wish', '想听', 'FF9800'),
    ],
    'games': [
        ('collect', '玩过', '4CAF50'),
        ('do', '在玩', '2196F3'),
        ('wish', '想玩', 'FF9800'),
    ],
}

# 共用样式
_THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

_HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
_HEADER_FILL = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center')

_ROW_EVEN_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
_ROW_ODD_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

_TITLE_FONT = Font(name='Microsoft YaHei', size=11)
_CELL_FONT = Font(name='Microsoft YaHei', size=10)
_LINK_FONT = Font(name='Microsoft YaHei', size=10, color='1565C0', underline='single')
_STAR_FONT = Font(name='Microsoft YaHei', size=10, color='FF8F00')


class DataStorage:
    def __init__(self):
        self.backup_dir = os.path.join(DATA_DIR, 'backup')
        os.makedirs(self.backup_dir, exist_ok=True)

    # ───────── JSON ─────────

    def save_json(self, data, filename):
        filepath = os.path.join(self.backup_dir, f"{filename}.json")
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  已保存: {filepath}")
        return filepath

    def save_movies_json(self, movies_data):
        return self.save_json(movies_data, 'movies')

    def save_books_json(self, books_data):
        return self.save_json(books_data, 'books')

    def save_all_json(self, all_data):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return self.save_json(all_data, f"douban_backup_{timestamp}")

    # ───────── Excel ─────────

    def save_excel(self, data, filename):
        filepath = os.path.join(self.backup_dir, f"{filename}.xlsx")
        wb = Workbook()
        # 删除默认 sheet
        wb.remove(wb.active)

        # 1) 总览 sheet
        self._write_overview_sheet(wb, data)

        # 2) 各类别 sheet
        for category in ['movies', 'books', 'music', 'games']:
            if category in data and data[category]:
                cat_cn, _ = _CATEGORY_CN[category]
                self._write_category_sheet(wb, cat_cn, category, data[category])

        # 确保至少有一个 sheet
        if len(wb.sheetnames) == 0:
            wb.create_sheet('空')

        wb.save(filepath)
        print(f"  已保存: {filepath}")
        return filepath

    def save_all_excel(self, all_data):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return self.save_excel(all_data, f"douban_backup_{timestamp}")

    # ───────── 总览 Sheet ─────────

    def _write_overview_sheet(self, wb, data):
        ws = wb.create_sheet('总览', 0)

        # 标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = '豆瓣数据备份报告'
        title_cell.font = Font(name='Microsoft YaHei', bold=True, size=18, color='37474F')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

        # 导出时间
        ws.merge_cells('A2:F2')
        time_cell = ws['A2']
        time_cell.value = f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        time_cell.font = Font(name='Microsoft YaHei', size=11, color='757575')
        time_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 22

        # 空行
        row = 4

        # 统计表头
        status_labels = self._get_overview_status_labels(data)
        headers = ['类别'] + status_labels + ['合计']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER
        ws.row_dimensions[row].height = 28

        # 统计数据行
        totals_by_col = [0] * len(status_labels)
        grand_total = 0

        for category in ['movies', 'books', 'music', 'games']:
            if category not in data or not data[category]:
                continue
            row += 1
            cat_cn, _ = _CATEGORY_CN[category]
            ws.cell(row=row, column=1, value=cat_cn).font = Font(name='Microsoft YaHei', bold=True, size=11)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).border = _THIN_BORDER

            cat_total = 0
            for si, (status_key, _, _) in enumerate(_STATUS_ORDER[category]):
                items = data[category].get(status_key, [])
                count = len(items)
                col = si + 2
                cell = ws.cell(row=row, column=col, value=count)
                cell.font = _CELL_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = _THIN_BORDER
                if si < len(totals_by_col):
                    totals_by_col[si] += count
                cat_total += count

            total_cell = ws.cell(row=row, column=len(headers), value=cat_total)
            total_cell.font = Font(name='Microsoft YaHei', bold=True, size=11)
            total_cell.alignment = Alignment(horizontal='center')
            total_cell.border = _THIN_BORDER
            grand_total += cat_total

            # 交替行色
            fill = _ROW_EVEN_FILL if (row % 2 == 0) else _ROW_ODD_FILL
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = fill

        # 合计行
        row += 1
        ws.cell(row=row, column=1, value='合计').font = Font(name='Microsoft YaHei', bold=True, size=11)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=1).border = _THIN_BORDER
        sum_fill = PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid')
        ws.cell(row=row, column=1).fill = sum_fill
        for si, t in enumerate(totals_by_col):
            cell = ws.cell(row=row, column=si + 2, value=t)
            cell.font = Font(name='Microsoft YaHei', bold=True, size=11)
            cell.alignment = Alignment(horizontal='center')
            cell.border = _THIN_BORDER
            cell.fill = sum_fill
        gt_cell = ws.cell(row=row, column=len(headers), value=grand_total)
        gt_cell.font = Font(name='Microsoft YaHei', bold=True, size=11, color='D32F2F')
        gt_cell.alignment = Alignment(horizontal='center')
        gt_cell.border = _THIN_BORDER
        gt_cell.fill = sum_fill

        # 列宽
        ws.column_dimensions['A'].width = 10
        for i in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

    def _get_overview_status_labels(self, data):
        """根据实际数据决定总览表的状态列标签"""
        # 找到第一个有数据的类别，取其状态标签
        for category in ['movies', 'books', 'music', 'games']:
            if category in data and data[category]:
                return [label for _, label, _ in _STATUS_ORDER[category]]
        return ['已完成', '进行中', '想要']

    # ───────── 类别 Sheet ─────────

    def _write_category_sheet(self, wb, sheet_name, category, cat_data):
        ws = wb.create_sheet(sheet_name)
        columns = _COLUMNS[category]
        col_count = len(columns)

        # 设置列宽
        for ci, (_, _, width) in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = width

        current_row = 1

        for status_key, status_label, color in _STATUS_ORDER[category]:
            items = cat_data.get(status_key, [])
            if not items:
                continue

            _, unit = _CATEGORY_CN[category]

            # ── 状态分组标题行 ──
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=col_count
            )
            group_cell = ws.cell(row=current_row, column=1)
            group_cell.value = f'  {status_label} ({len(items)}{unit})'
            group_cell.font = Font(name='Microsoft YaHei', bold=True, size=12, color='FFFFFF')
            group_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            group_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[current_row].height = 30
            current_row += 1

            # ── 表头行 ──
            for ci, (_, header, _) in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=ci, value=header)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGN
                cell.border = _THIN_BORDER
            ws.row_dimensions[current_row].height = 26
            header_row = current_row
            current_row += 1

            # ── 数据行 ──
            for idx, item in enumerate(items, 1):
                fill = _ROW_EVEN_FILL if (idx % 2 == 0) else _ROW_ODD_FILL

                for ci, (key, _, _) in enumerate(columns, 1):
                    if key == '_index':
                        value = idx
                        cell = ws.cell(row=current_row, column=ci, value=value)
                        cell.font = _CELL_FONT
                        cell.alignment = Alignment(horizontal='center')
                    elif key == '_link':
                        douban_id = item.get('douban_id', '')
                        item_type = item.get('type', category.rstrip('s'))
                        prefix = _LINK_PREFIX.get(item_type, 'https://www.douban.com/subject/')
                        if douban_id:
                            link_url = f'{prefix}{douban_id}/'
                            cell = ws.cell(row=current_row, column=ci)
                            cell.value = link_url
                            cell.hyperlink = link_url
                            cell.font = _LINK_FONT
                        else:
                            cell = ws.cell(row=current_row, column=ci, value='')
                            cell.font = _CELL_FONT
                    elif key == 'rating':
                        cell = ws.cell(row=current_row, column=ci, value=_rating_to_stars(item.get(key, '')))
                        cell.font = _STAR_FONT
                        cell.alignment = Alignment(horizontal='center')
                    elif key == 'title':
                        cell = ws.cell(row=current_row, column=ci, value=item.get(key, ''))
                        cell.font = _TITLE_FONT
                    else:
                        cell = ws.cell(row=current_row, column=ci, value=item.get(key, ''))
                        cell.font = _CELL_FONT

                    cell.fill = fill
                    cell.border = _THIN_BORDER
                    if key not in ('_index', 'rating', '_link', 'title'):
                        cell.alignment = Alignment(vertical='center', wrap_text=True)

                ws.row_dimensions[current_row].height = 22
                current_row += 1

            # 分组间空一行
            current_row += 1

        # 冻结首行（第一个表头行）
        # 找到第一个表头行位置
        ws.freeze_panes = 'A2'

    # ───────── 兼容旧接口 ─────────

    def get_backup_list(self):
        files = []
        for f in os.listdir(self.backup_dir):
            if f.endswith(('.json', '.xlsx')):
                filepath = os.path.join(self.backup_dir, f)
                files.append({
                    'name': f,
                    'path': filepath,
                    'size': os.path.getsize(filepath),
                    'modified': datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m-%d %H:%M:%S')
                })
        return sorted(files, key=lambda x: x['modified'], reverse=True)
