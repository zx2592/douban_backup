import json
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

import crawl_public
import storage


class ExcelSafetyTests(unittest.TestCase):
    def test_storage_save_json_wraps_data_with_metadata(self):
        data = {"movies": {"collect": [{"title": "Movie A"}]}}

        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            storage, "DATA_DIR", tmpdir
        ):
            data_storage = storage.DataStorage(
                metadata={
                    "backup_mode": "authenticated",
                    "selected_categories": ["movies"],
                    "user_id": "demo-user",
                }
            )
            path = data_storage.save_json(data, "with_metadata")

            with open(path, "r", encoding="utf-8") as file_obj:
                payload = json.load(file_obj)

        self.assertEqual(payload["data"], data)
        self.assertEqual(payload["metadata"]["backup_mode"], "authenticated")
        self.assertEqual(payload["metadata"]["selected_categories"], ["movies"])
        self.assertEqual(payload["metadata"]["user_id"], "demo-user")
        self.assertIn("generated_at", payload["metadata"])
        self.assertIn("timezone", payload["metadata"])

    def test_storage_save_excel_adds_metadata_sheet(self):
        data = {"movies": {"collect": []}}

        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            storage, "DATA_DIR", tmpdir
        ):
            data_storage = storage.DataStorage(
                metadata={
                    "backup_mode": "authenticated",
                    "selected_categories": ["movies"],
                }
            )
            path = data_storage.save_excel(data, "with_metadata_sheet")
            workbook = load_workbook(path)

        self.assertIn("元数据", workbook.sheetnames)
        worksheet = workbook["元数据"]
        rows = {
            worksheet[f"A{row}"].value: worksheet[f"B{row}"].value
            for row in range(1, worksheet.max_row + 1)
        }
        self.assertEqual(rows["backup_mode"], "authenticated")
        self.assertEqual(rows["selected_categories"], "movies")
        self.assertIn("generated_at", rows)
        self.assertIn("timezone", rows)

    def test_storage_save_excel_escapes_formula_like_values(self):
        data = {
            "movies": {
                "collect": [
                    {
                        "douban_id": "1",
                        "title": "=2+2",
                        "rating": "5",
                        "comment": "@cmd",
                        "tags": "-danger",
                        "date": "+2026-01-01",
                        "type": "movie",
                    }
                ]
            }
        }

        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            storage, "DATA_DIR", tmpdir
        ):
            data_storage = storage.DataStorage()
            path = data_storage.save_excel(data, "formula_safe")

            workbook = load_workbook(path)
            worksheet = workbook[workbook.sheetnames[-1]]

        self.assertEqual(worksheet["B3"].value, "'=2+2")
        self.assertEqual(worksheet["D3"].value, "'@cmd")
        self.assertEqual(worksheet["E3"].value, "'-danger")
        self.assertEqual(worksheet["F3"].value, "'+2026-01-01")

    def test_public_save_excel_escapes_formula_like_values(self):
        data = {
            "movies": {
                "collect": [
                    {
                        "douban_id": "1",
                        "title": "=SUM(1,1)",
                        "rating": "5",
                        "comment": "@comment",
                        "date": "-2026-01-01",
                        "cover": "+cover",
                    }
                ]
            }
        }

        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            crawl_public, "OUTPUT_DIR", tmpdir
        ):
            path = crawl_public.save_excel(data, "public_formula_safe")
            workbook = load_workbook(path)
            worksheet = workbook.active

        self.assertEqual(worksheet["C2"].value, "'=SUM(1,1)")
        self.assertEqual(worksheet["F2"].value, "'@comment")
        self.assertEqual(worksheet["G2"].value, "'-2026-01-01")
        self.assertEqual(worksheet["H2"].value, "'-2026-01-01")

    def test_public_save_json_wraps_data_with_metadata(self):
        data = {"movies": {"collect": []}}

        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            crawl_public, "OUTPUT_DIR", tmpdir
        ):
            path = crawl_public.save_json(
                data,
                "public_with_metadata",
                metadata={
                    "backup_mode": "public",
                    "selected_categories": ["movies"],
                    "user_id": "demo-user",
                },
            )

            with open(path, "r", encoding="utf-8") as file_obj:
                payload = json.load(file_obj)

        self.assertEqual(payload["data"], data)
        self.assertEqual(payload["metadata"]["backup_mode"], "public")
        self.assertEqual(payload["metadata"]["selected_categories"], ["movies"])
        self.assertEqual(payload["metadata"]["user_id"], "demo-user")
        self.assertIn("generated_at", payload["metadata"])
        self.assertIn("timezone", payload["metadata"])


class OverviewSheetTests(unittest.TestCase):
    HEADER_ROW = 4

    def _overview_headers(self, data):
        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            storage, "DATA_DIR", tmpdir
        ):
            data_storage = storage.DataStorage()
            path = data_storage.save_excel(data, "overview_headers")
            workbook = load_workbook(path)
            worksheet = workbook["总览"]
            return [
                worksheet.cell(row=self.HEADER_ROW, column=col).value
                for col in range(1, 6)
            ]

    def test_single_category_uses_its_own_status_labels(self):
        headers = self._overview_headers(
            {"books": {"collect": [{"title": "书"}], "reading": [], "wish": []}}
        )

        self.assertEqual(headers, ["类别", "已读", "在读", "想读", "合计"])

    def test_mixed_categories_use_neutral_status_labels(self):
        headers = self._overview_headers(
            {
                "movies": {"collect": [{"title": "片"}], "do": [], "wish": []},
                "books": {"collect": [{"title": "书"}], "reading": [], "wish": []},
            }
        )

        self.assertEqual(headers, ["类别", "已完成", "进行中", "想要", "合计"])


class CategorySheetTests(unittest.TestCase):
    def _category_sheet(self, data, sheet_name):
        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            storage, "DATA_DIR", tmpdir
        ):
            data_storage = storage.DataStorage()
            path = data_storage.save_excel(data, "category_sheet")
            return load_workbook(path)[sheet_name]

    def test_freeze_panes_keeps_first_header_row_visible(self):
        worksheet = self._category_sheet(
            {"movies": {"collect": [{"title": "片", "douban_id": "1"}]}},
            "电影",
        )

        self.assertEqual(worksheet["A2"].value, "序号")
        self.assertEqual(worksheet.freeze_panes, "A3")

    def test_freeze_panes_follows_first_non_empty_status_group(self):
        worksheet = self._category_sheet(
            {"movies": {"collect": [], "do": [{"title": "片", "douban_id": "1"}]}},
            "电影",
        )

        self.assertEqual(worksheet["A2"].value, "序号")
        self.assertEqual(worksheet.freeze_panes, "A3")


if __name__ == "__main__":
    unittest.main()
